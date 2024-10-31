import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

# Caminhos de exemplo
# O caminho para a planilha base, onde estão os dados originais
planilha_base = r'\\exemplo.xlsx'
# O caminho da pasta onde as planilhas separadas serão salvas
caminho_pasta_os_filial = r'\\exemplo_PART'

# Ler a planilha base (exemplo com dados fictícios)
# Supondo que a planilha contém colunas como 'Filial', 'Nome', 'Data', 'Valor'
df = pd.read_excel(planilha_base)

# Obter a lista de filiais únicas a partir da coluna 'Filial'
filial_unica = df['Filial'].unique()

# Processar cada filial
for filial in filial_unica:
    # Filtrar o DataFrame para a filial atual
    filial_df = df[df['Filial'] == filial]
    
    # Verificar e criar diretório para a filial, se necessário
    if not os.path.exists(caminho_pasta_os_filial):
        os.makedirs(caminho_pasta_os_filial)
    
    # Criar o arquivo Excel separado para a filial
    arquivo_filial = os.path.join(caminho_pasta_os_filial, f'{filial}.xlsx')
    filial_df.to_excel(arquivo_filial, index=False)
    
    # Carregar a planilha criada com openpyxl
    wb = load_workbook(arquivo_filial)
    ws = wb.active
    
    # Centralizar o conteúdo das células
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar automaticamente a largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Letra da coluna (A, B, C, etc.)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Adicionar margem ao comprimento e ajustar a largura
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Adicionar formatação de tabela
    tab = Table(displayName=f"Table_{filial}", ref=ws.dimensions)
    
    # Estilo da tabela
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Salvar o arquivo com a formatação aplicada
    wb.save(arquivo_filial)
    
print('Processo concluído e planilhas formatadas com colunas ajustadas para cada filial.')
